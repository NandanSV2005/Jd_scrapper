"""
Excel & JSON Writer Utility - handles exporting scraped job data to Excel files
and NextBuild-compatible JSON files in the scraped_jds folder.
"""

import os
import json
import re
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from scrapers.base_scraper import JobListing


COMMON_SKILLS_REF = [
    "React", "Next.js", "TypeScript", "JavaScript", "Node.js", "Python", "FastAPI",
    "Django", "Flask", "PostgreSQL", "MySQL", "MongoDB", "Redis", "Docker",
    "Kubernetes", "AWS", "GCP", "Azure", "GraphQL", "REST API", "Git",
    "PyTorch", "TensorFlow", "Java", "C++", "C#", "Go", "Rust", "TailwindCSS",
    "HTML", "CSS", "CI/CD", "Kafka", "Microservices", "System Design", "Swift", "iOS"
]

def extract_skills_from_text(text: str) -> List[str]:
    if not text:
        return ["React", "TypeScript", "Node.js", "Python", "PostgreSQL"]
    text_lower = text.lower()
    found = [s for s in COMMON_SKILLS_REF if s.lower() in text_lower]
    return found if found else ["React", "TypeScript", "Node.js", "Python", "REST API"]


class ExcelWriter:
    """Handles writing scraped job data to formatted Excel & JSON files."""

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    CELL_FONT = Font(name="Calibri", size=11)
    CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
    BULLET_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def _generate_filename(self, source: str = "") -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if source:
            clean_source = "".join(c for c in source if c.isalnum() or c in " _-").strip()
            return f"jobs_{clean_source}_{timestamp}.xlsx"
        return f"jobs_scraped_{timestamp}.xlsx"

    def write_jobs(self, jobs: List[JobListing], filename: str = "") -> str:
        """Write job listings to an Excel file with proper formatting."""
        if not filename:
            sources = set(j.source for j in jobs)
            source_str = "_".join(filter(None, sources)) if sources else ""
            filename = self._generate_filename(source_str)

        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Job Listings"

        col_widths = {
            "A": 8,    # S.No
            "B": 30,   # Company
            "C": 35,   # Job Role
            "D": 80,   # Job Description
            "E": 15,   # Source
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        headers = ["S.No", "Company Name", "Job Role", "Job Description (Bullet Points)", "Source"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        ws.freeze_panes = "A2"

        for row_idx, job in enumerate(jobs, 2):
            if job.description_bullets:
                bullets_text = "\n".join(f"• {bullet}" for bullet in job.description_bullets)
            elif job.description:
                bullets_text = f"• {job.description}"
            else:
                bullets_text = "No description available"

            row_data = [
                row_idx - 1,
                job.company,
                job.job_role,
                bullets_text,
                job.source,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.CELL_FONT
                cell.border = self.THIN_BORDER

                if col_idx in (2, 3, 5):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                elif col_idx == 4:
                    cell.alignment = self.BULLET_ALIGNMENT
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="top")

            if row_idx % 2 == 0:
                for col_idx in range(1, 6):
                    ws.cell(row=row_idx, column=col_idx).fill = self.ALT_ROW_FILL

        wb.save(filepath)
        print(f"[Excel] Saved {len(jobs)} job listings to: {filepath}")
        
        # Also export to NextBuild scraped_jds folder automatically
        try:
            self.export_to_nextbuild_folder(jobs)
        except Exception as e:
            print(f"[NextBuild Export Info]: {e}")

        return filepath

    def export_to_nextbuild_folder(self, jobs: List[JobListing], target_dir: str = "../scraped_jds") -> List[str]:
        """Export scraped jobs into NextBuild's scraped_jds/*.json format."""
        saved_files = []
        os.makedirs(target_dir, exist_ok=True)

        for job in jobs:
            clean_company = re.sub(r"[^a-zA-Z0-9]+", "_", job.company.lower()).strip("_") or "scraped"
            clean_role = re.sub(r"[^a-zA-Z0-9]+", "_", job.job_role.lower()).strip("_") or "job"
            filename = f"{clean_company}_{clean_role}.json"
            filepath = os.path.join(target_dir, filename)

            desc = job.description or ("\n".join(job.description_bullets) if job.description_bullets else f"{job.job_role} position at {job.company}.")
            skills = extract_skills_from_text(desc)

            payload = {
                "id": f"scraped-{clean_company}-{clean_role}",
                "company": job.company or "Featured Organization",
                "title": job.job_role or "Software Engineer",
                "location": "Scraped Target Location",
                "url": "#",
                "descriptionSnippet": desc[:220] + ("..." if len(desc) > 220 else ""),
                "requiredSkills": skills,
                "fullDescription": desc,
            }

            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2)

            saved_files.append(filepath)
            print(f"[NextBuild Scraped JD] Exported: {filepath}")

        return saved_files
