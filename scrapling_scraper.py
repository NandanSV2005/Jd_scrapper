"""
Job Scraper Pro — Scrapling & Playwright Naukri Engine (v4.5)
Extracts job data from Naukri listing search pages AND single job posting URLs.
Supports both card parsing and full single job page parsing.
"""

import re
import sys
import time
import argparse
from typing import Optional

from bs4 import BeautifulSoup


# ─── Helpers ────────────────────────────────────────────────────────────

def extract_bullets(text: str) -> list:
    """Extract bullet points from description text."""
    if not text or len(text) < 15:
        return []
    lines = re.split(r'[•·●◆◇▪▸▹►▻‣⁃⦿✦✧]\s*|[\n\r]+|(?:\d+[.)])\s*', text)
    result = []
    for line in lines:
        line = re.sub(r'\s+', ' ', line).strip()
        if len(line) > 10:
            result.append(line)
            if len(result) >= 50:
                break
    return result


def clean_text(text: str) -> str:
    """Clean and normalize text."""
    if not text:
        return ''
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def export_excel(jobs: list, filename: str = "jd_scraper_output.xlsx") -> str:
    """Export scraped jobs to a styled Excel file."""
    if not jobs:
        print("[!] No jobs to export.")
        return ""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[!] openpyxl not installed.")
        return ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    data_font = Font(name='Calibri', size=10)
    data_align = Alignment(vertical='top', wrap_text=True)

    headers = [
        'S.No', 'Company Name', 'Job Role',
        'Job Description', 'Key Skills',
        'Location', 'Experience',
        'Source'
    ]
    col_widths = [6, 28, 35, 70, 40, 22, 15, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, job in enumerate(jobs, 2):
        desc_bullets = job.get('descriptionBullets', [])
        desc_text = '\n'.join(f'• {b}' for b in desc_bullets) if desc_bullets else job.get('description', '')
        skills_str = ', '.join(job.get('skills', [])) if job.get('skills') else ''

        row_data = [
            row_idx - 1,
            job.get('company', ''),
            job.get('role', ''),
            desc_text,
            skills_str,
            job.get('location', ''),
            job.get('experience', ''),
            job.get('source', 'Naukri'),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    wb.save(filename)
    print(f"[OK] Saved {len(jobs)} jobs to '{filename}'")
    return filename


class NaukriScraper:
    """Scrapes Naukri.com listings & single job postings."""

    def __init__(self, headless: bool = True, verbose: bool = True):
        self.headless = headless
        self.verbose = verbose
        self._logs = []

    def log(self, msg: str):
        if self.verbose:
            print(msg)
        self._logs.append(msg)

    def _fetch_with_playwright(self, url: str) -> str:
        """Fetch HTML using Playwright as primary/fallback stealth fetcher."""
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=self.headless)
                context = browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                    viewport={"width": 1920, "height": 1080},
                )
                page = context.new_page()
                self.log(f"  Playwright fetching: {url}")
                page.goto(url, wait_until="domcontentloaded", timeout=45000)
                time.sleep(3)
                html = page.content()
                browser.close()
                return html
        except Exception as e:
            self.log(f"  [Playwright fetch error]: {e}")
            return ""

    def _fetch_listing(self, url: str) -> str:
        """Fetch page HTML using Scrapling DynamicFetcher with Playwright fallback."""
        try:
            from scrapling.fetchers import DynamicFetcher
            self.log("  Using DynamicFetcher (stealth browser)...")
            page = DynamicFetcher.fetch(url, headless=self.headless, network_idle=True)
            if page and hasattr(page, 'html_content') and len(page.html_content) > 1000:
                return page.html_content
        except Exception as e:
            self.log(f"  [Scrapling info]: {e}")

        return self._fetch_with_playwright(url)

    def extract_single_job(self, soup: BeautifulSoup, url: str) -> Optional[dict]:
        """Extract job details from a single Naukri job posting page."""
        title = ""
        for sel in [".jd-header-title", "h1[class*='title']", "h1.styles_j24-wrapper__title", "h1"]:
            el = soup.select_one(sel)
            if el and el.get_text().strip():
                title = clean_text(el.get_text())
                break

        company = ""
        for sel in [".jd-header-title-company", "a[class*='company']", "[class*='comp-name']", "a[href*='careers']"]:
            el = soup.select_one(sel)
            if el and el.get_text().strip():
                company = clean_text(el.get_text())
                break

        description = ""
        for sel in [".job-details-description", ".jd-desc", "div[class*='styles_j24-wrapper__description']", "section[class*='description']", "div[class*='job-desc']"]:
            el = soup.select_one(sel)
            if el and el.get_text().strip():
                description = clean_text(el.get_text(separator="\n"))
                break

        skills = []
        for sel in [".key-skill", ".skill", "[class*='chip']", "[class*='key-skill']"]:
            els = soup.select(sel)
            if els:
                skills = [clean_text(e.get_text()) for e in els if clean_text(e.get_text())]
                if skills:
                    break

        if not title and not company and not description:
            return None

        bullets = extract_bullets(description)
        return {
            'company': company or 'Naukri Hiring Company',
            'role': title or 'Software Developer',
            'description': description or 'Naukri job description posting.',
            'descriptionBullets': bullets or [description] if description else [],
            'skills': skills,
            'location': 'India / Remote',
            'experience': '0-5 Yrs',
            'source': 'Naukri',
        }

    def extract_cards(self, html: str) -> list:
        """Parse job cards from listing page HTML using BeautifulSoup."""
        soup = BeautifulSoup(html, 'lxml')

        # Check if single job page first
        single_job = self.extract_single_job(soup, "")
        if single_job:
            self.log("  [OK] Extracted single job page details!")
            return [single_job]

        # Find job card containers
        cards = soup.select('div.cust-job-tuple, div[class*="cust-job-tuple"], div.sjw__tuple, article[class*="job"], div[class*="jobTuple"], div.srp-jobtuple-wrapper')
        if not cards:
            cards = soup.find_all("div", attrs={"data-job-id": re.compile(r".*")})

        if not cards:
            self.log("[!] No job cards found via CSS selectors")
            return []

        self.log(f"  [OK] Found {len(cards)} job cards")
        jobs = []
        for card in cards:
            try:
                job = self._parse_card(card)
                if job.get('role'):
                    jobs.append(job)
            except Exception:
                continue

        self.log(f"  [OK] Parsed {len(jobs)} jobs from cards")
        return jobs

    def _parse_card(self, card) -> dict:
        """Extract data from a single job card element."""
        title_el = card.select_one('a[href*="job-listings"], a[class*="title"], a.title')
        role = clean_text(title_el.get_text()) if title_el else ''

        company_el = card.select_one('a[href*="-jobs-careers-"], a[class*="company"], .comp-name')
        company = clean_text(company_el.get_text()) if company_el else ''

        all_text = card.get_text(separator=' ', strip=True)
        all_text = re.sub(r'\s+', ' ', all_text)

        exp_match = re.search(r'(\d+[-\s]to\s*\d+|\d+[-\s]*\d+)\s*Yrs?', all_text, re.I)
        experience = exp_match.group(1).strip() + ' Yrs' if exp_match else ''

        locations = ['Bengaluru', 'Bangalore', 'Mumbai', 'Delhi', 'Pune', 'Hyderabad',
                     'Chennai', 'Kolkata', 'Ahmedabad', 'Gurugram', 'Gurgaon', 'Noida',
                     'Remote', 'Work From Home', 'India']
        location = ''
        for loc in locations:
            if loc.lower() in all_text.lower():
                location = loc
                break

        skill_elements = card.select('[class*="skill"], [class*="key"], a[class*="skill"], span[class*="tag"]')
        skills = [clean_text(se.get_text()) for se in skill_elements if clean_text(se.get_text())]

        desc_segment = clean_text(all_text[:280])
        bullets = extract_bullets(desc_segment)

        return {
            'company': company or 'Naukri Organization',
            'role': role or 'Software Role',
            'description': desc_segment,
            'descriptionBullets': bullets,
            'skills': skills[:10],
            'location': location,
            'experience': experience,
            'source': 'Naukri',
        }

    def scrape_listing(self, url: str, max_jobs: int = 50) -> list:
        """Scrape Naukri page."""
        self.log(f"\n{'='*60}\nNAUKRI SCRAPER v4.5\n{'='*60}\nURL: {url}\n")
        html = self._fetch_listing(url)
        if not html or len(html) < 500:
            self.log("[FAIL] Page fetch failed or empty")
            return []

        all_jobs = self.extract_cards(html)
        if not all_jobs:
            self.log("[FAIL] No jobs extracted")
            return []

        if len(all_jobs) > max_jobs:
            all_jobs = all_jobs[:max_jobs]

        self.log(f"RESULT: {len(all_jobs)} jobs extracted!")
        return all_jobs

    def export_excel(self, jobs: list, filename: str = "jd_scraper_output.xlsx") -> str:
        return export_excel(jobs, filename)


def main():
    p = argparse.ArgumentParser(description='Job Scraper Pro — Naukri Engine')
    p.add_argument('url', help='Naukri job search or listing URL')
    p.add_argument('-o', '--output', default='jd_scraper_output.xlsx')
    p.add_argument('-m', '--max', type=int, default=50)
    args = p.parse_args()

    scraper = NaukriScraper()
    jobs = scraper.scrape_listing(args.url, max_jobs=args.max)
    if jobs:
        scraper.export_excel(jobs, args.output)
        print(f"\nDone! {len(jobs)} jobs -> '{args.output}'")
        return 0
    return 1


if __name__ == '__main__':
    sys.exit(main())
